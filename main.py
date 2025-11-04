import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import sys
import os
import parametros


def validar_data(data):
    """
    Valida se uma data é válida verificando:
    - Mês entre 1 e 12
    - Dia válido para o mês (considerando anos bissextos)
    - Ausência de valores negativos
    - Ano razoável (entre 1900 e 2100)
    """
    if not isinstance(data, datetime):
        return False

    if data.year < 0 or data.month < 0 or data.day < 0:
        return False

    if data.month < 1 or data.month > 12:
        return False

    if data.year < 1900 or data.year > 2100:
        return False

    try:
        from datetime import date
        teste = date(data.year, data.month, data.day)
        return True
    except ValueError:
        return False


def determinar_periodo_dre(workbook):
    """
    Determina o período inicial e final da DRE baseado nos dados das abas
    Vendas, Custo_Despesas e Folha.
    """
    datas_encontradas = []
    datas_invalidas = []

    def converter_data(valor):
        """Converte um valor para datetime, lidando com diferentes formatos."""
        if valor is None:
            return None
        try:
            if isinstance(valor, datetime):
                return valor
            elif isinstance(valor, (int, float)):
                if valor < 0 or valor > 1000000:
                    return None
                try:
                    from datetime import date, timedelta
                    excel_epoch = date(1899, 12, 30)
                    data = excel_epoch + timedelta(days=int(valor))
                    return datetime(data.year, data.month, data.day)
                except (ValueError, OverflowError):
                    return None
            elif isinstance(valor, str):
                valor = valor.strip()
                if not valor:
                    return None
                formatos = ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d %H:%M:%S',
                            '%d-%m-%Y', '%m-%d-%Y']
                for fmt in formatos:
                    try:
                        return datetime.strptime(valor, fmt)
                    except ValueError:
                        continue
                return None
            else:
                return None
        except (ValueError, TypeError, AttributeError):
            return None

    if 'Vendas' in workbook.sheetnames:
        ws_vendas = workbook['Vendas']
        for row_num, row in enumerate(ws_vendas.iter_rows(min_row=2, min_col=6, max_col=6, values_only=False), start=2):
            if row[0].value is not None and row[0].value not in ['', None]:
                data = converter_data(row[0].value)
                if data:
                    if validar_data(data):
                        datas_encontradas.append(data)
                    else:
                        datas_invalidas.append(('Vendas', row_num, row[0].value))
                else:
                    datas_invalidas.append(('Vendas', row_num, row[0].value))

    if 'Custo_Despesas' in workbook.sheetnames:
        ws_custo = workbook['Custo_Despesas']
        for row_num, row in enumerate(ws_custo.iter_rows(min_row=2, min_col=3, max_col=3, values_only=False), start=2):
            if row[0].value is not None and row[0].value not in ['', None]:
                data = converter_data(row[0].value)
                if data:
                    if validar_data(data):
                        datas_encontradas.append(data)
                    else:
                        datas_invalidas.append(('Custo_Despesas', row_num, row[0].value))
                else:
                    datas_invalidas.append(('Custo_Despesas', row_num, row[0].value))

    if 'Folha' in workbook.sheetnames:
        ws_folha = workbook['Folha']
        for row_num, row in enumerate(ws_folha.iter_rows(min_row=2, min_col=1, max_col=1, values_only=False), start=2):
            if row[0].value is not None and row[0].value not in ['', None]:
                data = converter_data(row[0].value)
                if data:
                    if validar_data(data):
                        datas_encontradas.append(data)
                    else:
                        datas_invalidas.append(('Folha', row_num, row[0].value))
                else:
                    datas_invalidas.append(('Folha', row_num, row[0].value))

    if not datas_encontradas:
        print("⚠ Aviso: Não foram encontradas datas válidas nas planilhas. Usando valores padrão.")
        return '2024-01-01', 12, datas_invalidas

    data_min = min(datas_encontradas)
    data_max = max(datas_encontradas)

    data_inicial = datetime(data_min.year, data_min.month, 1)
    data_final = datetime(data_max.year, data_max.month, 1)

    meses_diff = (data_final.year - data_inicial.year) * 12 + (data_final.month - data_inicial.month)
    num_meses = meses_diff + 1

    data_inicial_str = data_inicial.strftime('%Y-%m-%d')
    print(f"✓ Período detectado: {data_inicial_str} a {data_final.strftime('%Y-%m-%d')} ({num_meses} meses)")

    return data_inicial_str, num_meses, datas_invalidas


def verificar_abas_fontes(workbook, abas_necessarias):
    abas_existentes = workbook.sheetnames
    abas_faltantes = [aba for aba in abas_necessarias if aba not in abas_existentes]
    if abas_faltantes:
        raise ValueError(f"ERRO: As seguintes abas fontes não foram encontradas: {', '.join(abas_faltantes)}")
    print(f"✓ Todas as abas fontes foram encontradas: {', '.join(abas_necessarias)}")


def criar_aba_dre_se_nao_existir(workbook):
    if 'DRE' in workbook.sheetnames:
        print("✓ Aba DRE já existe. Será reconstruída.")
        workbook.remove(workbook['DRE'])
    ws_dre = workbook.create_sheet('DRE', 0)
    print("✓ Aba DRE criada com sucesso.")
    return ws_dre


def configurar_cabecalho_dre(ws_dre, data_inicial='2024-01-01', num_meses=12):
    ws_dre['A1'] = 'DRE'
    ws_dre['A1'].font = Font(size=14, bold=True)
    ws_dre.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws_dre['A1'].alignment = Alignment(horizontal='left')

    data_base = datetime.strptime(data_inicial, '%Y-%m-%d')
    ws_dre.cell(row=3, column=4).value = data_base
    ws_dre.cell(row=3, column=4).number_format = 'mm/yy'
    ws_dre.cell(row=3, column=4).alignment = Alignment(horizontal='center')

    for i in range(1, num_meses):
        col_letra = get_column_letter(4 + i)
        col_anterior = get_column_letter(4 + i - 1)
        ws_dre[f'{col_letra}3'] = f'=EDATE({col_anterior}3, 1)'
        ws_dre[f'{col_letra}3'].number_format = 'mm/yy'
        ws_dre[f'{col_letra}3'].alignment = Alignment(horizontal='center')


def aplicar_formatacao_dre(ws_dre, num_colunas=12):
    formato_milhares = '#,##0.00,_);(#,##0.00,); -'
    formato_porcentagem = '0.0%'
    font_italico = Font(italic=True)
    font_negrito = Font(bold=True)
    alinhamento_num = Alignment(horizontal='right')
    alinhamento_texto = Alignment(horizontal='left')
    fill_azul = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
    fill_branco = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    ws_dre.column_dimensions['A'].width = 2.5
    ws_dre.column_dimensions['B'].width = 2.5
    ws_dre.column_dimensions['C'].width = 22

    for col_off in range(1, 4 + num_colunas):
        cell = ws_dre.cell(row=1, column=col_off)
        cell.fill = fill_azul

    for j in range(num_colunas):
        c = ws_dre.cell(row=3, column=4 + j)
        c.number_format = 'mm/yy'
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill = fill_branco

    max_lin = 56
    for i in range(2, max_lin + 1):
        for j in range(1, 4 + num_colunas):
            cell = ws_dre.cell(row=i, column=j)
            if cell.fill != fill_azul:
                cell.fill = fill_branco

    termos_percentagem = ["%", "Growth", "% da Receita"]

    for col in [1, 2, 3]:
        for i in range(4, max_lin):
            cell = ws_dre.cell(row=i, column=col)
            if cell.value is not None:
                cell_value_str = str(cell.value)
                is_percentagem = any(termo in cell_value_str for termo in termos_percentagem)
                if is_percentagem:
                    cell.font = Font(bold=False, italic=True)
                elif col == 2 and "Margem" in cell_value_str:
                    cell.font = Font(bold=True, italic=True)
                elif col == 2 or col == 3:
                    cell.font = font_negrito
                cell.alignment = alinhamento_texto

    linhas_percentuais = [5, 8, 10, 12, 14, 17, 22, 24, 26, 28, 31, 36, 41, 52]
    for i in range(4, max_lin):
        for j in range(num_colunas):
            c = ws_dre.cell(row=i, column=4 + j)
            if i in linhas_percentuais or (
                    ws_dre.cell(row=i, column=2).value and "Margem" in str(ws_dre.cell(row=i, column=2).value)):
                c.number_format = formato_porcentagem
                c.font = font_italico
            else:
                c.number_format = formato_milhares
                c.font = Font(italic=False)
            c.alignment = alinhamento_num


def construir_estrutura_dre(workbook, ws_dre, num_colunas=12):
    linha = 4
    ws_dre.cell(row=linha, column=2).value = 'Receita '
    for i in range(num_colunas):
        col = 4 + i
        col_letra = get_column_letter(col)
        formula = f'=SUMIFS(Vendas!$E:$E,Vendas!$F:$F,">="&EOMONTH(DRE!{col_letra}$3,-1)+1,Vendas!$F:$F,"<="&EOMONTH(DRE!{col_letra}$3,0))'
        ws_dre.cell(row=linha, column=col).value = formula

    linha = 5
    ws_dre.cell(row=linha, column=2).value = 'Growth %'
    for i in range(num_colunas):
        col = 4 + i
        if i == 0:
            ws_dre.cell(row=linha, column=col).value = None
        else:
            col_letra = get_column_letter(col)
            col_anterior = get_column_letter(col - 1)
            formula = f'=({col_letra}4/{col_anterior}4)-1'
            ws_dre.cell(row=linha, column=col).value = formula

    linha = 7
    ws_dre.cell(row=linha, column=2).value = 'CMV (-)'
    for i in range(num_colunas):
        col = 4 + i
        col_letra = get_column_letter(col)
        formula = f'=-SUM({col_letra}9, {col_letra}11, {col_letra}13)'
        ws_dre.cell(row=linha, column=col).value = formula

    linha = 8
    ws_dre.cell(row=linha, column=2).value = '% da Receita'
    for i in range(num_colunas):
        col = 4 + i
        col_letra = get_column_letter(col)
        formula = f'={col_letra}7/{col_letra}4'
        ws_dre.cell(row=linha, column=col).value = formula

    linha = 9
    ws_dre.cell(row=linha, column=3).value = 'Armazenagem'
    for i in range(num_colunas):
        col = 4 + i
        col_letra = get_column_letter(col)
        formula = f'=SUMIFS(Custo_Despesas!$B:$B,Custo_Despesas!$C:$C,">="&EOMONTH(DRE!{col_letra}$3,-1)+1,Custo_Despesas!$C:$C,"<="&EOMONTH(DRE!{col_letra}$3,0),Custo_Despesas!$A:$A,$C9)'
        ws_dre.cell(row=linha, column=col).value = formula

    linha = 10
    ws_dre.cell(row=linha, column=3).value = '% da Receita'
    for i in range(num_colunas):
        col = 4 + i
        col_letra = get_column_letter(col)
        formula = f'={col_letra}9/{col_letra}4'
        ws_dre.cell(row=linha, column=col).value = formula

    linha = 11
    ws_dre.cell(row=linha, column=3).value = 'Frete'
    for i in range(num_colunas):
        col = 4 + i
        col_letra = get_column_letter(col)
        formula = f'=SUMIFS(Custo_Despesas!$B:$B,Custo_Despesas!$C:$C,">="&EOMONTH(DRE!{col_letra}$3,-1)+1,Custo_Despesas!$C:$C,"<="&EOMONTH(DRE!{col_letra}$3,0),Custo_Despesas!$A:$A,$C11)'
        ws_dre.cell(row=linha, column=col).value = formula

    linha = 12
    ws_dre.cell(row=linha, column=3).value = '% da Receita'
    for i in range(num_colunas):
        col = 4 + i
        col_letra = get_column_letter(col)
        formula = f'={col_letra}11/{col_letra}4'
        ws_dre.cell(row=linha, column=col).value = formula

    linha = 13
    ws_dre.cell(row=linha, column=3).value = 'Matéria-prima'
    for i in range(num_colunas):
        col = 4 + i
        col_letra = get_column_letter(col)
        formula = f'=SUMIFS(Custo_Despesas!$B:$B,Custo_Despesas!$C:$C,">="&EOMONTH(DRE!{col_letra}$3,-1)+1,Custo_Despesas!$C:$C,"<="&EOMONTH(DRE!{col_letra}$3,0),Custo_Despesas!$A:$A,$C13)'
        ws_dre.cell(row=linha, column=col).value = formula

    linha = 14
    ws_dre.cell(row=linha, column=3).value = '% da Receita'
    for i in range(num_colunas):
        col = 4 + i
        col_letra = get_column_letter(col)
        formula = f'={col_letra}13/{col_letra}4'
        ws_dre.cell(row=linha, column=col).value = formula

    linha = 16
    ws_dre.cell(row=linha, column=2).value = 'Lucro bruto'
    for i in range(num_colunas):
        col = 4 + i
        col_letra = get_column_letter(col)
        formula = f'={col_letra}4+{col_letra}7'
        ws_dre.cell(row=linha, column=col).value = formula

    linha = 17
    ws_dre.cell(row=linha, column=2).value = 'Margem Bruta %'
    for i in range(num_colunas):
        col = 4 + i
        col_letra = get_column_letter(col)
        formula = f'={col_letra}16/{col_letra}4'
        ws_dre.cell(row=linha, column=col).value = formula

    linha = 19
    ws_dre.cell(row=linha, column=2).value = 'SG&A (-)'
    for i in range(num_colunas):
        col = 4 + i
        col_letra = get_column_letter(col)
        formula = f'=-SUM({col_letra}21, {col_letra}23, {col_letra}25, {col_letra}27)'
        ws_dre.cell(row=linha, column=col).value = formula

    linha = 21
    ws_dre.cell(row=linha, column=3).value = 'Marketing'
    for i in range(num_colunas):
        col = 4 + i
        col_letra = get_column_letter(col)
        formula = f'=SUMIFS(Custo_Despesas!$B:$B,Custo_Despesas!$C:$C,">="&EOMONTH(DRE!{col_letra}$3,-1)+1,Custo_Despesas!$C:$C,"<="&EOMONTH(DRE!{col_letra}$3,0),Custo_Despesas!$A:$A,$C21)'
        ws_dre.cell(row=linha, column=col).value = formula

    linha = 22
    ws_dre.cell(row=linha, column=3).value = '% da Receita'
    for i in range(num_colunas):
        col = 4 + i
        col_letra = get_column_letter(col)
        formula = f'={col_letra}21/{col_letra}4'
        ws_dre.cell(row=linha, column=col).value = formula

    linha = 23
    ws_dre.cell(row=linha, column=3).value = 'Comercial'
    for i in range(num_colunas):
        col = 4 + i
        col_letra = get_column_letter(col)
        formula = f'=SUMIFS(Custo_Despesas!$B:$B,Custo_Despesas!$C:$C,">="&EOMONTH(DRE!{col_letra}$3,-1)+1,Custo_Despesas!$C:$C,"<="&EOMONTH(DRE!{col_letra}$3,0),Custo_Despesas!$A:$A,$C23)'
        ws_dre.cell(row=linha, column=col).value = formula

    linha = 24
    ws_dre.cell(row=linha, column=3).value = '% da Receita'
    for i in range(num_colunas):
        col = 4 + i
        col_letra = get_column_letter(col)
        formula = f'={col_letra}23/{col_letra}4'
        ws_dre.cell(row=linha, column=col).value = formula

    linha = 25
    ws_dre.cell(row=linha, column=3).value = 'Administrativo'
    for i in range(num_colunas):
        col = 4 + i
        col_letra = get_column_letter(col)
        formula = f'=SUMIFS(Custo_Despesas!$B:$B,Custo_Despesas!$C:$C,">="&EOMONTH(DRE!{col_letra}$3,-1)+1,Custo_Despesas!$C:$C,"<="&EOMONTH(DRE!{col_letra}$3,0),Custo_Despesas!$A:$A,$C25)'
        ws_dre.cell(row=linha, column=col).value = formula

    linha = 26
    ws_dre.cell(row=linha, column=3).value = '% da Receita'
    for i in range(num_colunas):
        col = 4 + i
        col_letra = get_column_letter(col)
        formula = f'={col_letra}25/{col_letra}4'
        ws_dre.cell(row=linha, column=col).value = formula

    linha = 27
    ws_dre.cell(row=linha, column=3).value = 'Folha'
    for i in range(num_colunas):
        col_letra = get_column_letter(4 + i)
        formula = (
            f'=SUMIFS(Folha!$C:$C, Folha!$A:$A, ">= " & EOMONTH(DRE!{col_letra}$3, -1) + 1, Folha!$A:$A, "<= " & EOMONTH(DRE!{col_letra}$3, 0))'
            f'+SUMIFS(Folha!$D:$D, Folha!$A:$A, ">= " & EOMONTH(DRE!{col_letra}$3, -1) + 1, Folha!$A:$A, "<= " & EOMONTH(DRE!{col_letra}$3, 0))'
            f'+SUMIFS(Folha!$E:$E, Folha!$A:$A, ">= " & EOMONTH(DRE!{col_letra}$3, -1) + 1, Folha!$A:$A, "<= " & EOMONTH(DRE!{col_letra}$3, 0))'
        )
        ws_dre.cell(row=linha, column=4 + i).value = formula

    linha = 28
    ws_dre.cell(row=linha, column=3).value = '% da Receita'
    for i in range(num_colunas):
        col = 4 + i
        col_letra = get_column_letter(col)
        formula = f'={col_letra}27/{col_letra}4'
        ws_dre.cell(row=linha, column=col).value = formula

    linha = 30
    ws_dre.cell(row=linha, column=2).value = 'EBITDA'
    for i in range(num_colunas):
        col = 4 + i
        col_letra = get_column_letter(col)
        formula = f'={col_letra}4+{col_letra}7+{col_letra}19'
        ws_dre.cell(row=linha, column=col).value = formula

    linha = 31
    ws_dre.cell(row=linha, column=2).value = 'Margem EBITDA %'
    for i in range(num_colunas):
        col = 4 + i
        col_letra = get_column_letter(col)
        formula = f'={col_letra}30/{col_letra}4'
        ws_dre.cell(row=linha, column=col).value = formula

    linha = 33
    ws_dre.cell(row=linha, column=2).value = 'D&A (-)'
    for i in range(num_colunas):
        col = 4 + i
        col_letra_inv = get_column_letter(10 + i)
        formula = f'=-Investimentos!{col_letra_inv}1'
        ws_dre.cell(row=linha, column=col).value = formula

    linha = 35
    ws_dre.cell(row=linha, column=2).value = 'EBIT'
    for i in range(num_colunas):
        col = 4 + i
        col_letra = get_column_letter(col)
        formula = f'={col_letra}30+{col_letra}33'
        ws_dre.cell(row=linha, column=col).value = formula

    linha = 36
    ws_dre.cell(row=linha, column=2).value = 'Margem Operacional %'
    for i in range(num_colunas):
        col = 4 + i
        col_letra = get_column_letter(col)
        formula = f'={col_letra}35/{col_letra}4'
        ws_dre.cell(row=linha, column=col).value = formula

    # --- Juros (-) ---
    linha = 38
    ws_dre.cell(row=linha, column=2).value = 'Juros (-)'
    if 'Financiamento' in workbook.sheetnames:
        for i in range(num_colunas):
            col = 4 + i
            col_letra = get_column_letter(col)
            # This formula dynamically finds the interest for the correct month.
            # It assumes interest values are in row 5 and dates are in row 4 of the 'Financiamento' sheet.
            formula = (
                f'=-SUMIFS(Financiamento!$J$5:$AZ$5, '
                f'Financiamento!$J$4:$AZ$4, ">="&EOMONTH(DRE!{col_letra}$3,-1)+1, '
                f'Financiamento!$J$4:$AZ$4, "<="&EOMONTH(DRE!{col_letra}$3,0))'
            )
            ws_dre.cell(row=linha, column=col).value = formula
    else:
        print("⚠ Aviso: Aba 'Financiamento' não encontrada. Juros (-) permanecerão zerados.")
        for i in range(num_colunas):
            ws_dre.cell(row=linha, column=4 + i).value = 0

    linha = 40
    ws_dre.cell(row=linha, column=2).value = 'EBT'
    for i in range(num_colunas):
        col = 4 + i
        col_letra = get_column_letter(col)
        formula = f'={col_letra}35+{col_letra}38'
        ws_dre.cell(row=linha, column=col).value = formula

    linha = 41
    ws_dre.cell(row=linha, column=2).value = '% da Receita'
    for i in range(num_colunas):
        col = 4 + i
        col_letra = get_column_letter(col)
        formula = f'={col_letra}40/{col_letra}4'
        ws_dre.cell(row=linha, column=col).value = formula

    for row in range(43, 51):
        ws_dre.row_dimensions[row].outline_level = 1
    ws_dre.row_dimensions[50].collapsed = True
    ws_dre.sheet_properties.outline_summary_below = True

    linha = 43
    ws_dre.cell(row=linha, column=2).value = 'Prejuizo Acumulado *'
    linha = 44
    ws_dre.cell(row=linha, column=3).value = 'Inicio'
    ws_dre.cell(row=linha, column=4).value = 0
    for i in range(1, num_colunas):
        col = 4 + i
        col_anterior = get_column_letter(col - 1)
        formula = f'={col_anterior}47'
        ws_dre.cell(row=linha, column=col).value = formula

    linha = 45
    ws_dre.cell(row=linha, column=3).value = 'Saldo Adquirido'
    for i in range(num_colunas):
        col = 4 + i
        col_letra = get_column_letter(col)
        formula = f'=-IF({col_letra}40<0,-{col_letra}40,0)'
        ws_dre.cell(row=linha, column=col).value = formula

    linha = 46
    ws_dre.cell(row=linha, column=3).value = 'Saldo Utilizado'
    for i in range(num_colunas):
        col = 4 + i
        col_letra = get_column_letter(col)
        formula = f'=IF({col_letra}$40>0, MIN({col_letra}$40*30%, -{col_letra}$44), 0)'
        ws_dre.cell(row=linha, column=col).value = formula

    linha = 47
    ws_dre.cell(row=linha, column=3).value = 'Final'
    for i in range(num_colunas):
        col = 4 + i
        formula = f'=SUM({get_column_letter(col)}44:{get_column_letter(col)}46)'
        ws_dre.cell(row=linha, column=col).value = formula

    linha = 49
    ws_dre.cell(row=linha, column=2).value = 'Base de calculo '
    for i in range(num_colunas):
        col = 4 + i
        col_letra = get_column_letter(col)
        formula = f'=IF({col_letra}40<0,0,{col_letra}40-{col_letra}46)'
        ws_dre.cell(row=linha, column=col).value = formula

    linha = 51
    ws_dre.cell(row=linha, column=2).value = f'Impostos (-)'
    for i in range(num_colunas):
        col = 4 + i
        col_letra = get_column_letter(col)
        taxa_decimal = parametros.taxa_imposto / 100
        formula = f'=-{col_letra}49*{taxa_decimal}'
        ws_dre.cell(row=linha, column=col).value = formula

    linha = 52
    ws_dre.cell(row=linha, column=2).value = 'Taxa Efetiva de Imposto %'
    for i in range(num_colunas):
        col = 4 + i
        col_letra = get_column_letter(col)
        formula = f'=ABS({col_letra}51)/{col_letra}35'
        ws_dre.cell(row=linha, column=col).value = formula

    linha = 54
    ws_dre.cell(row=linha, column=2).value = 'Lucro (prejuízo) líquido'
    for i in range(num_colunas):
        col = 4 + i
        col_letra = get_column_letter(col)
        formula = f'=SUM({col_letra}51,{col_letra}40)'
        ws_dre.cell(row=linha, column=col).value = formula

def calcular_waterfall_depreciacao(workbook, data_inicial, num_meses):
    """
    Calcula o waterfall de depreciação (D&A) APENAS para o período da DRE.
    Depreciação começa no MÊS SEGUINTE após o investimento ser lançado.
    Usa EDATE como na DRE para garantir sincronização perfeita das datas.
    Vida útil dos ativos vem de parametros.py
    """
    if 'Investimentos' not in workbook.sheetnames:
        print("⚠ Aviso: Aba 'Investimentos' não encontrada. Ignorando cálculo de D&A.")
        return

    ws_inv = workbook['Investimentos']

    # Usar vida_util_ativos de parametros.py
    investments = []
    row = 2
    while ws_inv.cell(row=row, column=1).value is not None:
        date = ws_inv.cell(row=row, column=1).value
        desc = ws_inv.cell(row=row, column=2).value
        value = ws_inv.cell(row=row, column=3).value

        if isinstance(date, datetime) and desc and value:
            # Busca a vida útil em parametros, usa padrão se não encontrar
            vida_util = parametros.vida_util_ativos.get(desc, parametros.vida_util_padrao)
            investments.append({
                'date': date,
                'description': desc,
                'value': value,
                'vida_util': vida_util
            })
        row += 1

    print(f"✓ {len(investments)} investimento(s) encontrado(s)")

    waterfall_start_col = 5  # Coluna E
    waterfall_start_row = 3

    max_row = 1
    for row in range(1, ws_inv.max_row + 1):
        if (ws_inv.cell(row=row, column=1).value is not None or
                ws_inv.cell(row=row, column=2).value is not None or
                ws_inv.cell(row=row, column=3).value is not None):
            max_row = row

    for row in range(1, waterfall_start_row + max_row):
        for column in range(waterfall_start_col, waterfall_start_col + 100):
            try:
                cell = ws_inv.cell(row=row, column=column)
                cell.font = Font(bold=False, color='000000')
                cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                cell.value = None
            except AttributeError:
                continue  # Proteção extra

    ws_inv.cell(row=waterfall_start_row, column=waterfall_start_col).value = 'WATERFALL DE DEPRECIAÇÃO'
    ws_inv.cell(row=waterfall_start_row, column=waterfall_start_col).font = Font(size=12, bold=True)

    waterfall_header_row = waterfall_start_row + 2

    headers = ['Ativo', 'Descrição', 'Valor', 'Vida Útil (anos)', 'Deprec. Mensal']
    for col_offset, header in enumerate(headers):
        col = waterfall_start_col + col_offset
        cell = ws_inv.cell(row=waterfall_header_row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # USAR EDATE como na DRE
    date_base = datetime.strptime(data_inicial, '%Y-%m-%d')

    # Primeira coluna de data
    col_primeira_data = waterfall_start_col + 5
    ws_inv.cell(row=waterfall_header_row, column=col_primeira_data).value = date_base
    ws_inv.cell(row=waterfall_header_row, column=col_primeira_data).number_format = 'mm/yy'

    # Replicar com EDATE como na DRE
    for i in range(1, num_meses):
        col_letra_atual = get_column_letter(col_primeira_data + i)
        col_letra_anterior = get_column_letter(col_primeira_data + i - 1)

        cell = ws_inv.cell(row=waterfall_header_row, column=col_primeira_data + i)
        cell.value = f'=EDATE({col_letra_anterior}{waterfall_header_row},1)'
        cell.number_format = 'mm/yy'
        cell.alignment = Alignment(horizontal='center')

    waterfall_data_start_row = waterfall_header_row + 1
    asset_counter = 1

    for inv_idx, inv in enumerate(investments):
        row_num = waterfall_data_start_row + inv_idx

        col_ativo = waterfall_start_col
        col_desc = waterfall_start_col + 1
        col_valor = waterfall_start_col + 2
        col_vida_util = waterfall_start_col + 3
        col_deprec_mensal = waterfall_start_col + 4

        ws_inv.cell(row=row_num, column=col_ativo).value = asset_counter
        ws_inv.cell(row=row_num, column=col_desc).value = inv['description']
        ws_inv.cell(row=row_num, column=col_valor).value = inv['value']
        ws_inv.cell(row=row_num, column=col_valor).number_format = '#,##0.00'
        ws_inv.cell(row=row_num, column=col_vida_util).value = inv['vida_util']

        col_valor_letra = get_column_letter(col_valor)
        col_vida_util_letra = get_column_letter(col_vida_util)
        formula_deprec = f'={col_valor_letra}{row_num}/({col_vida_util_letra}{row_num}*12)'
        ws_inv.cell(row=row_num, column=col_deprec_mensal).value = formula_deprec
        ws_inv.cell(row=row_num, column=col_deprec_mensal).number_format = '#,##0.00'

        # Data de INÍCIO da DEPRECIAÇÃO = MÊS SEGUINTE ao investimento
        data_investimento = inv['date'].replace(day=1)
        mes_deprec_inicio = data_investimento.month + 1
        ano_deprec_inicio = data_investimento.year
        if mes_deprec_inicio > 12:
            mes_deprec_inicio = 1
            ano_deprec_inicio += 1
        data_inicio_deprec = data_investimento.replace(year=ano_deprec_inicio, month=mes_deprec_inicio)

        # Data final da depreciação (vida útil * 12 meses depois do início)
        n_meses_total = inv['vida_util'] * 12

        mes_final = data_inicio_deprec.month + n_meses_total
        ano_final = data_inicio_deprec.year + (mes_final - 1) // 12
        mes_final = ((mes_final - 1) % 12) + 1
        data_fim_deprec = data_inicio_deprec.replace(year=ano_final, month=mes_final)

        # Preencher depreciação por mês - APENAS para o período da DRE
        for col_idx in range(num_meses):
            col = col_primeira_data + col_idx
            col_letra = get_column_letter(col)

            col_deprec_letra = get_column_letter(col_deprec_mensal)

            # Fórmula: se data da coluna >= data inicio E < data fim, então deprecia
            formula = (
                f'=IF(AND({col_letra}{waterfall_header_row}>='
                f'DATE({data_inicio_deprec.year},{data_inicio_deprec.month},1),'
                f'{col_letra}{waterfall_header_row}<'
                f'DATE({data_fim_deprec.year},{data_fim_deprec.month},1)),'
                f'{col_deprec_letra}{row_num},0)'
            )

            ws_inv.cell(row=row_num, column=col).value = formula
            ws_inv.cell(row=row_num, column=col).number_format = '#,##0.00'

        asset_counter += 1

    total_row = waterfall_data_start_row + len(investments) + 1
    col_total = waterfall_start_col

    ws_inv.cell(row=1, column=col_total).value = 'TOTAL D&A'
    ws_inv.cell(row=1, column=col_total).font = Font(bold=True, color='FFFFFF')
    ws_inv.cell(row=1, column=col_total).fill = PatternFill(start_color='000000', end_color='000000',
                                                                    fill_type='solid')

    for col_idx in range(num_meses):
        col = col_primeira_data + col_idx
        col_letra = get_column_letter(col)

        primeira_linha = waterfall_data_start_row
        ultima_linha = total_row - 1

        formula = f'=SUM({col_letra}{primeira_linha}:{col_letra}{ultima_linha})'
        cell = ws_inv.cell(row=1, column=col)
        cell.value = formula
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
        cell.number_format = '#,##0.00'

    print(f"✓ Waterfall de depreciação calculado com sucesso")
    print(f"  - Total de ativos: {len(investments)}")
    print(f"  - Vida útil configurada em parametros.py")

def converter_periodo_especifico(inicio_str, final_str):
    try:
        inicio = datetime.strptime(inicio_str, '%m/%y')
        final = datetime.strptime(final_str, '%m/%y')

        data_inicial = datetime(inicio.year, inicio.month, 1)
        data_final = datetime(final.year, final.month, 1)

        meses_diff = (data_final.year - data_inicial.year) * 12 + (data_final.month - data_inicial.month)
        num_meses = meses_diff + 1

        data_inicial_str = data_inicial.strftime('%Y-%m-%d')

        return data_inicial_str, num_meses

    except ValueError as e:
        raise ValueError(f"Erro ao converter período: {e}. Formato esperado: MM/YY (ex: '01/24')")


def automatizar_dre(caminho_arquivo='entrada.xlsx', data_inicial=None, num_meses=None):
    try:
        abas_necessarias = ['Vendas', 'Custo_Despesas', 'Folha', 'Investimentos', 'Financiamento']

        print("=" * 80)
        print("AUTOMATIZAÇÃO DA ABA DRE")
        print("=" * 80)

        print(f"\nCarregando arquivo: {caminho_arquivo}")
        wb = openpyxl.load_workbook(caminho_arquivo)
        print(f"✓ Arquivo carregado com sucesso.")

        datas_invalidas = []

        if data_inicial is None or num_meses is None:
            if parametros.auto_detectar_periodo:
                print(f"\nDeterminando período automaticamente...")
                data_inicial, num_meses, datas_invalidas = determinar_periodo_dre(wb)
            else:
                print(f"\nUsando período específico dos parâmetros...")
                print(f" Período: {parametros.periodo_inicio} a {parametros.periodo_final}")
                data_inicial, num_meses = converter_periodo_especifico(
                    parametros.periodo_inicio,
                    parametros.periodo_final
                )

        print(f"✓ Período configurado: {data_inicial} ({num_meses} meses)")

        print(f"\nVerificando datas inválidas nas planilhas...")

        _, _, datas_invalidas = determinar_periodo_dre(wb)

        print(f"\nVerificando abas fontes...")
        verificar_abas_fontes(wb, abas_necessarias)

        print(f"\nCriando aba DRE...")
        ws_dre = criar_aba_dre_se_nao_existir(wb)

        print(f"\nConfigurando cabeçalho...")
        configurar_cabecalho_dre(ws_dre, data_inicial, num_meses)

        print(f"\nConstruindo estrutura da DRE...")
        construir_estrutura_dre(wb, ws_dre, num_meses)

        print(f"\nCalculando waterfall de depreciação...")
        calcular_waterfall_depreciacao(wb, data_inicial, num_meses)

        print(f"\nAplicando formatação automática...")
        aplicar_formatacao_dre(ws_dre, num_meses)

        print(f"\nAjustando freeze panes...")
        ws_dre.freeze_panes = 'D4'

        print(f"\nSalvando arquivo...")
        wb.save(caminho_arquivo)

        print(f"✓ Arquivo salvo com sucesso em: {caminho_arquivo}")

        print("\n" + "=" * 80)
        print("DRE CONSTRUÍDA COM SUCESSO!")
        print("=" * 80)

        print(f"\nAbrindo planilha...")
        caminho_absoluto = os.path.abspath(caminho_arquivo)
        os.startfile(caminho_absoluto)
        print(f"✓ Planilha aberta com sucesso.")

        if datas_invalidas:
            print(f"\n❌ ERRO: datas invalidas encontradas:")
            for aba, linha, valor in datas_invalidas:
                print(f"   - Aba '{aba}', Linha {linha}: {valor}")
            print(f"   Total: {len(datas_invalidas)} data(s) inválida(s) foram ignoradas.")

    except ValueError as ve:
        print(f"\n❌ ERRO: {ve}")
    except Exception as e:
        print(f"\n❌ ERRO INESPERADO: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    automatizar_dre()
